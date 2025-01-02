import openpyxl
from S3costcalculatins import s3_cost_assisments
from datatransfer import Data_transfer
from awsQcostcalculation import Sqs_cost_assesments
from vpcCostCalculation import VPC_cost_assessment
from reportGeneration import ReportGenerating
from lambdaCostcalculation import Lambda_cost_assessments
from directConnect import DirectConnect_cost_assessment

# print("test")

path =r"C:\Users\smaram\Desktop\Book1.xlsx"

ws = openpyxl.load_workbook(path)

workbook = ws.active

direcconnect = DirectConnect_cost_assessment(workbook=workbook)
direcconnect.process()


# print(ws.sheetnames)

# for sheet in ws.sheetnames:
#     paramountdataTranfer = Data_transfer(ws[sheet]) #* AWS direct connect is included as data transfer
#     paramountdataTranfer.process()

#     if sheet  == 'AmazonVPC':
#         # continue
#         try:
#             paramountVPC = VPC_cost_assessment(ws[sheet])
#             paramountVPC.process()
#         except Exception as e:
#             print(f"error {e} at {sheet}")

#     if sheet == 'AmazonS3':
#         try:
#             paramountS3 = s3_cost_assisments(ws[sheet])
#             paramountS3.process()
#         except Exception as e:
#             print(f"error {e} at {sheet}")

#     if sheet == 'AWSQueueService':
#         try:
#         # continue
#             paramountQ = Sqs_cost_assesments(ws[sheet])
#             paramountQ.process()
#         except Exception as e:
#             print(f"error {e} at {sheet}")
            
#     if sheet == 'AWSLambda':
#         try:
#             paramountLambda = Lambda_cost_assessments(ws[sheet])
#             paramountLambda.process()
#         except Exception as e:
#             print(f"error {e} at {sheet}")

ws.save(path)